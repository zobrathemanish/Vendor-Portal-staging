from flask import Flask, render_template, request, redirect, url_for, flash, send_from_directory, session
import os
import json
from datetime import datetime
from werkzeug.utils import secure_filename
from azure.storage.blob import BlobServiceClient
import hashlib
from openpyxl import Workbook


# ---------------------------------------
# CONFIGURATION
# ---------------------------------------
app = Flask(__name__)
app.secret_key = "fgi_vendor_portal_secret"   #change later
app.config['SESSION_TYPE'] = 'filesystem'

# Local upload path (Phase 1 temp before Azure)
UPLOAD_FOLDER = os.path.join(os.getcwd(), "uploads")
TEMPLATE_FOLDER = os.path.join(os.getcwd(), "data", "templates")
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
ALLOWED_EXTENSIONS = {'xml', 'xlsx'}


# Azure Storage
AZURE_CONNECTION_STRING = os.getenv("AZURE_STORAGE_CONNECTION_STRING")
AZURE_CONTAINER_NAME = "bronze"

if not AZURE_CONNECTION_STRING:
    print("⚠ WARNING: AZURE_STORAGE_CONNECTION_STRING is not set. "
          "Azure uploads will fail until you configure it.")


# ---------------------------------------
# HELPER FUNCTIONS
# ---------------------------------------

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def save_file(file, vendor_name, subfolder):
    """Save file locally (Phase 1 temp)."""
    filename = secure_filename(file.filename)
    vendor_folder = os.path.join(app.config['UPLOAD_FOLDER'], vendor_name, subfolder)
    os.makedirs(vendor_folder, exist_ok=True)
    filepath = os.path.join(vendor_folder, filename)
    file.save(filepath)
    return filepath


def compute_file_hash(filepath):
    """Compute SHA-256 of ZIP without loading into RAM."""
    sha = hashlib.sha256()
    with open(filepath, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            sha.update(chunk)
    return sha.hexdigest()


def get_blob_service_client():
    if not AZURE_CONNECTION_STRING:
        raise RuntimeError("AZURE_STORAGE_CONNECTION_STRING is not configured.")
    return BlobServiceClient.from_connection_string(AZURE_CONNECTION_STRING)


def get_latest_asset_hash(container_client, vendor_folder):
    """Return last uploaded asset hash from Azure manifest."""
    prefix = f"raw/{vendor_folder}/logs/"
    blobs = list(container_client.list_blobs(name_starts_with=prefix))

    if not blobs:
        return None  # No manifest present yet

    blobs_sorted = sorted(blobs, key=lambda b: b.name, reverse=True)
    latest_manifest_blob = blobs_sorted[0]

    manifest_data = container_client.download_blob(latest_manifest_blob.name).readall()
    manifest = json.loads(manifest_data)

    return manifest.get("assets_hash")


def delete_old_asset_zips(container_client, vendor_folder):
    prefix = f"raw/{vendor_folder}/assets/"
    blobs = list(container_client.list_blobs(name_starts_with=prefix))

    if len(blobs) <= 1:
        return  # Nothing to delete

    blobs_sorted = sorted(blobs, key=lambda b: b.name)

    for old_blob in blobs_sorted[:-1]:
        container_client.delete_blob(old_blob.name)
        print(f"[CLEANUP] Deleted old asset ZIP: {old_blob.name}")


def upload_blob(local_path, blob_path):
    """Upload a file to Azure under the bronze container."""
    blob_service_client = get_blob_service_client()
    container_client = blob_service_client.get_container_client(AZURE_CONTAINER_NAME)
    blob_client = container_client.get_blob_client(blob_path)

    with open(local_path, "rb") as data:
        blob_client.upload_blob(data, overwrite=True)

    print(f"✅ Uploaded to Azure: {blob_path}")
    return f"{AZURE_CONTAINER_NAME}/{blob_path}"


def upload_to_azure_bronze(vendor, xml_local_path, pricing_local_path, zip_path):
    """
    Upload product XML, pricing XLSX, and assets ZIP to Azure Bronze.
    Includes: hash check, skip-if-same, smart deletion, manifest updates.
    """
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    vendor_folder = f"vendor={vendor}"

    blob_service_client = get_blob_service_client()
    container_client = blob_service_client.get_container_client(AZURE_CONTAINER_NAME)

    # --------------- XML + Pricing always uploaded ---------------
    xml_blob_path = f"raw/{vendor_folder}/product/{timestamp}_product.xml"
    pricing_blob_path = f"raw/{vendor_folder}/pricing/{timestamp}_pricing.xlsx"

    xml_blob_full = upload_blob(xml_local_path, xml_blob_path)
    pricing_blob_full = upload_blob(pricing_local_path, pricing_blob_path)

    # --------------- Assets ZIP handling ---------------
    assets_blob_full = None
    assets_hash = None

    if zip_path and os.path.isfile(zip_path):

        # Compute hash of new ZIP
        new_hash = compute_file_hash(zip_path)

        # Get previous asset hash from last manifest
        last_hash = get_latest_asset_hash(container_client, vendor_folder)

        if last_hash == new_hash:
            print("🟡 Assets unchanged. Skipping ZIP upload.")
        else:
            print("🟢 Assets changed. Uploading new ZIP...")

            assets_hash = new_hash
            assets_blob_path = f"raw/{vendor_folder}/assets/{timestamp}_assets.zip"
            assets_blob_full = upload_blob(zip_path, assets_blob_path)

            # Clean up old ZIPs
            delete_old_asset_zips(container_client, vendor_folder)
    else:
        print("⚠ No ZIP path provided or file does not exist. Skipping assets upload.")

    # --------------- Create manifest ---------------
    manifest = {
        "vendor": vendor,
        "timestamp": timestamp,
        "azure_xml_blob": xml_blob_full,
        "azure_pricing_blob": pricing_blob_full,
        "azure_assets_blob": assets_blob_full,
        "assets_hash": assets_hash,
    }

    local_manifest_dir = os.path.join(app.config['UPLOAD_FOLDER'], vendor, "opticat")
    os.makedirs(local_manifest_dir, exist_ok=True)
    local_manifest_path = os.path.join(local_manifest_dir, f"manifest_{timestamp}.json")

    with open(local_manifest_path, "w") as f:
        json.dump(manifest, f, indent=4)

    manifest_blob_path = f"raw/{vendor_folder}/logs/manifest_{timestamp}.json"
    upload_blob(local_manifest_path, manifest_blob_path)

    print(f"📄 Manifest created and uploaded: {manifest_blob_path}")

def create_multi_product_excel(
    item_rows,
    desc_rows,
    ext_rows,
    attr_rows,
    interchange_rows,
    package_rows,
    asset_rows,
    price_rows,
    output_dir
):
    """
    Create an Excel file with separate tabs for:
      - Item_Master
      - Descriptions
      - Extended_Info
      - Attributes
      - Part_Interchange
      - Packages
      - Digital_Assets
      - Pricing

    Each *_rows argument is a list of dicts where keys match the column names.
    """
    os.makedirs(output_dir, exist_ok=True)

    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    filename = f"single_products_batch_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)

    wb = Workbook()

    # 1) Item_Master
    ws = wb.active
    ws.title = "Item_Master"
    item_cols = [
        "Vendor",
        "Part Number",
        "UNSPSC",
        "HazmatFlag",
        "Product Status",
        "Barcode Type",
        "Barcode Number",
        "Quantity UOM",
        "Quantity Size",
        "VMRS Code",
    ]
    ws.append(item_cols)
    for row in item_rows:
        ws.append([
            row.get("Vendor", ""),
            row.get("Part Number", ""),
            row.get("UNSPSC", ""),
            row.get("HazmatFlag", ""),
            row.get("Product Status", ""),
            row.get("Barcode Type", ""),
            row.get("Barcode Number", ""),
            row.get("Quantity UOM", ""),
            row.get("Quantity Size", ""),
            row.get("VMRS Code", ""),
        ])

    # 2) Descriptions
    ws = wb.create_sheet("Descriptions")
    desc_cols = [
        "SKU",
        "Description Change Type",
        "Description Code",
        "Description Value",
        "Sequence",
    ]
    ws.append(desc_cols)
    for row in desc_rows:
        ws.append([
            row.get("SKU", ""),
            row.get("Description Change Type", ""),
            row.get("Description Code", ""),
            row.get("Description Value", ""),
            row.get("Sequence", ""),
        ])

    # 3) Extended_Info
    ws = wb.create_sheet("Extended_Info")
    ext_cols = [
        "SKU",
        "Extended Info Change Type",
        "Extended Info Code",
        "Extended Info Value",
    ]
    ws.append(ext_cols)
    for row in ext_rows:
        ws.append([
            row.get("SKU", ""),
            row.get("Extended Info Change Type", ""),
            row.get("Extended Info Code", ""),
            row.get("Extended Info Value", ""),
        ])

    # 4) Attributes
    ws = wb.create_sheet("Attributes")
    attr_cols = [
        "SKU",
        "Attribute Change Type",
        "Attribute Name",
        "Attribute Value",
    ]
    ws.append(attr_cols)
    for row in attr_rows:
        ws.append([
            row.get("SKU", ""),
            row.get("Attribute Change Type", ""),
            row.get("Attribute Name", ""),
            row.get("Attribute Value", ""),
        ])

    # 5) Part_Interchange
    ws = wb.create_sheet("Part_Interchange")
    interchange_cols = [
        "SKU",
        "Part Interchange Change Type",
        "Brand Label",
        "Part Number",
    ]
    ws.append(interchange_cols)
    for row in interchange_rows:
        ws.append([
            row.get("SKU", ""),
            row.get("Part Interchange Change Type", ""),
            row.get("Brand Label", ""),
            row.get("Part Number", ""),
        ])

    # 6) Packages
    ws = wb.create_sheet("Packages")
    package_cols = [
        "SKU",
        "Package Change Type",
        "Package UOM",
        "Package Quantity of Eaches",
        "Weight UOM",
        "Weight",
    ]
    ws.append(package_cols)
    for row in package_rows:
        ws.append([
            row.get("SKU", ""),
            row.get("Package Change Type", ""),
            row.get("Package UOM", ""),
            row.get("Package Quantity of Eaches", ""),
            row.get("Weight UOM", ""),
            row.get("Weight", ""),
        ])

    # 7) Digital_Assets
    ws = wb.create_sheet("Digital_Assets")
    asset_cols = [
        "SKU",
        "Digital Change Type",
        "Media Type",
        "FileName",
        "FileLocalPath",
    ]
    ws.append(asset_cols)
    for row in asset_rows:
        ws.append([
            row.get("SKU", ""),
            row.get("Digital Change Type", ""),
            row.get("Media Type", ""),
            row.get("FileName", ""),
            row.get("FileLocalPath", ""),
        ])

    # 8) Pricing
    ws = wb.create_sheet("Pricing")
    price_cols = [
        "Vendor",
        "Part Number",
        "Pricing Method",
        "Currency",
        "MOQ Unit",
        "MOQ",
        "Pricing Change Type",
        "Pricing Type",
        "List Price",
        "Jobber Price",
        "Discount %",
        "Multiplier",
        "Pricing Amount",
        "Tier Min Qty",
        "Tier Max Qty",
        "Effective Date",
        "Start Date",
        "End Date",
        "Core Part Number",
        "Core Cost",
        "Notes",
    ]
    ws.append(price_cols)
    for row in price_rows:
        ws.append([
            row.get("Vendor", ""),
            row.get("Part Number", ""),
            row.get("Pricing Method", ""),
            row.get("Currency", ""),
            row.get("MOQ Unit", ""),
            row.get("MOQ", ""),
            row.get("Pricing Change Type", ""),
            row.get("Pricing Type", ""),
            row.get("List Price", ""),
            row.get("Jobber Price", ""),
            row.get("Discount %", ""),
            row.get("Multiplier", ""),
            row.get("Pricing Amount", ""),
            row.get("Tier Min Qty", ""),
            row.get("Tier Max Qty", ""),
            row.get("Effective Date", ""),
            row.get("Start Date", ""),
            row.get("End Date", ""),
            row.get("Core Part Number", ""),
            row.get("Core Cost", ""),
            row.get("Notes", ""),
        ])

    wb.save(filepath)
    return filepath



# def upload_single_product_excel_to_azure(vendor_name, local_path):
#     timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
#     vendor_folder = f"vendor={vendor_name}"
#     blob_path = f"raw/{vendor_folder}/manual/{timestamp}_single_product.xlsx"
#     return upload_blob(local_path, blob_path)

def validate_single_product_new(form: dict) -> tuple[bool, list[str]]:

    errors = []

    vendor = form.get("vendor_name", "").strip()
    sku = form.get("sku", "").strip()
    unspsc = form.get("unspsc_code", "").strip()
    hazmat = form.get("hazmat_flag", "").strip()
    status = form.get("product_status", "").strip()
    barcode_type = form.get("barcode_type", "").strip()
    barcode_number = form.get("barcode_number", "").strip()
    quantity_uom = form.get("quantity_uom", "").strip()
    quantity_size = form.get("quantity_size", "").strip()
    vmrs = form.get("vmrs_code", "").strip()

    # Required basic
    if not vendor:
        errors.append("Vendor is required.")
    elif vendor not in VENDOR_LIST:
        errors.append(f"Vendor '{vendor}' is not in the allowed list.")

    if not sku:
        errors.append("Part Number (SKU) is required.")

    # UNSPSC
    if unspsc:
        if not unspsc.isdigit() or len(unspsc) != 8:
            errors.append("UNSPSC Code must be exactly 8 numeric digits if provided.")

    # Hazmat
    if hazmat and hazmat not in ["Y", "N"]:
        errors.append("Hazardous Material must be Y or N.")

    # Status
    if status and status not in PRODUCT_STATUS:
        errors.append("Product Status must be one of: " + ", ".join(PRODUCT_STATUS))

    # Barcode
    if barcode_number:
        if not barcode_number.isdigit():
            errors.append("Barcode Number must be numeric.")
        else:
            if barcode_type == "UPC" and len(barcode_number) != 12:
                errors.append("UPC Barcode must be exactly 12 digits.")
            elif barcode_type == "EAN" and len(barcode_number) != 14:
                errors.append("EAN Barcode must be exactly 14 digits.")
            elif not barcode_type and len(barcode_number) not in (12, 14):
                errors.append("Barcode must be 12-digit UPC or 14-digit EAN.")

    if barcode_type and barcode_type not in ["UPC", "EAN"]:
        errors.append("Barcode Type must be UPC or EAN.")

    # Quantity / UOM
    if quantity_uom and quantity_uom not in QUANTITY_UOM:
        errors.append("Quantity Size Unit must be one of: " + ", ".join(QUANTITY_UOM))

    if quantity_size:
        try:
            float(quantity_size)
        except ValueError:
            errors.append("Quantity Size must be numeric.")

    # --- Descriptions: check lengths for DES/SHO ---
    desc_codes = form.getlist("desc_code[]")
    desc_values = form.getlist("desc_value[]")

    for code, value in zip(desc_codes, desc_values):
        code = (code or "").strip()
        value = (value or "").strip()
        if code in ("DES", "SHO") and value and len(value) > 40:
            errors.append(f"Description with code {code} must be <= 40 characters.")

    # --- MULTI-LEVEL PRICING VALIDATION ---
    level_types = form.getlist("level_type[]")
    level_price_change_types = form.getlist("level_price_change_type[]")
    level_moq_uoms = form.getlist("level_moq_uom[]")
    level_moq_qtys = form.getlist("level_moq_qty[]")
    level_currencies = form.getlist("level_currency[]")
    level_methods = form.getlist("level_pricing_method[]")

    # Method-specific arrays
    net_list = form.getlist("level_net_list_price[]")
    net_costs = form.getlist("level_net_net_cost[]")
    net_effective_dates = form.getlist("level_net_effective_date[]")

    pl_list = form.getlist("level_pl_list_price[]")
    pl_jobber = form.getlist("level_pl_jobber_price[]")
    pl_net = form.getlist("level_pl_net_cost[]")
    pl_effective_dates = form.getlist("level_pl_effective_date[]")

    db_base = form.getlist("level_db_base_price[]")
    db_discount = form.getlist("level_db_discount_pct[]")
    db_list_opt = form.getlist("level_db_list_price_opt[]")
    db_effective_dates = form.getlist("level_db_effective_date[]")

    ehc_base = form.getlist("level_ehc_base_price[]")
    ehc_cb = form.getlist("level_ehc_canadian_blue[]")
    ehc_qty_case = form.getlist("level_ehc_qty_case[]")
    ehc_upc_each = form.getlist("level_ehc_upc_each[]")
    ehc_upc_case = form.getlist("level_ehc_upc_case[]")
    ehc_moq = form.getlist("level_ehc_moq[]")

    ehc_abmbsk_each = form.getlist("level_ehc_abmbsk_each[]")
    ehc_abmbsk_case = form.getlist("level_ehc_abmbsk_case[]")
    ehc_bc_each = form.getlist("level_ehc_bc_each[]")
    ehc_bc_case = form.getlist("level_ehc_bc_case[]")
    ehc_nl_each = form.getlist("level_ehc_nl_each[]")
    ehc_nl_case = form.getlist("level_ehc_nl_case[]")
    ehc_ns_each = form.getlist("level_ehc_ns_each[]")
    ehc_ns_case = form.getlist("level_ehc_ns_case[]")
    ehc_nbqc_each = form.getlist("level_ehc_nbqc_each[]")
    ehc_nbqc_case = form.getlist("level_ehc_nbqc_case[]")
    ehc_pei_each = form.getlist("level_ehc_pei_each[]")
    ehc_pei_case = form.getlist("level_ehc_pei_case[]")
    ehc_yk_each = form.getlist("level_ehc_yk_each[]")
    ehc_yk_case = form.getlist("level_ehc_yk_case[]")

    pr_price = form.getlist("level_pr_promo_price[]")
    pr_start = form.getlist("level_pr_start_date[]")
    pr_end = form.getlist("level_pr_end_date[]")

    qt_price = form.getlist("level_qt_price[]")
    qt_number = form.getlist("level_qt_number[]")
    qt_start = form.getlist("level_qt_start_date[]")
    qt_end = form.getlist("level_qt_end_date[]")

    td_price = form.getlist("level_td_price[]")
    td_number = form.getlist("level_td_number[]")
    td_start = form.getlist("level_td_start_date[]")
    td_end = form.getlist("level_td_end_date[]")

    level_tier_min_qtys = form.getlist("level_tier_min_qty[]")
    level_tier_max_qtys = form.getlist("level_tier_max_qty[]")

    n_levels = len(level_types)

    if n_levels == 0:
        errors.append("At least one pricing level is required.")
        return (len(errors) == 0, errors)

    def pad(lst):
        return lst + [""] * (n_levels - len(lst))

    # Make sure all arrays have same length to avoid IndexError
    level_price_change_types = pad(level_price_change_types)
    level_moq_uoms = pad(level_moq_uoms)
    level_moq_qtys = pad(level_moq_qtys)
    level_currencies = pad(level_currencies)
    level_methods = pad(level_methods)
    net_list = pad(net_list)
    net_costs = pad(net_costs)
    net_effective_dates = pad(net_effective_dates)
    pl_list = pad(pl_list)
    pl_jobber = pad(pl_jobber)
    pl_net = pad(pl_net)
    pl_effective_dates = pad(pl_effective_dates)
    db_base = pad(db_base)
    db_discount = pad(db_discount)
    db_list_opt = pad(db_list_opt)
    db_effective_dates = pad(db_effective_dates)
    ehc_base = pad(ehc_base)
    ehc_cb = pad(ehc_cb)
    ehc_qty_case = pad(ehc_qty_case)
    ehc_upc_each = pad(ehc_upc_each)
    ehc_upc_case = pad(ehc_upc_case)
    ehc_moq = pad(ehc_moq)
    ehc_abmbsk_each = pad(ehc_abmbsk_each)
    ehc_abmbsk_case = pad(ehc_abmbsk_case)
    ehc_bc_each = pad(ehc_bc_each)
    ehc_bc_case = pad(ehc_bc_case)
    ehc_nl_each = pad(ehc_nl_each)
    ehc_nl_case = pad(ehc_nl_case)
    ehc_ns_each = pad(ehc_ns_each)
    ehc_ns_case = pad(ehc_ns_case)
    ehc_nbqc_each = pad(ehc_nbqc_each)
    ehc_nbqc_case = pad(ehc_nbqc_case)
    ehc_pei_each = pad(ehc_pei_each)
    ehc_pei_case = pad(ehc_pei_case)
    ehc_yk_each = pad(ehc_yk_each)
    ehc_yk_case = pad(ehc_yk_case)
    pr_price = pad(pr_price)
    pr_start = pad(pr_start)
    pr_end = pad(pr_end)
    qt_price = pad(qt_price)
    qt_number = pad(qt_number)
    qt_start = pad(qt_start)
    qt_end = pad(qt_end)
    td_price = pad(td_price)
    td_number = pad(td_number)
    td_start = pad(td_start)
    td_end = pad(td_end)
    level_tier_min_qtys = pad(level_tier_min_qtys)
    level_tier_max_qtys = pad(level_tier_max_qtys)

    any_level_used = False
    today_str = datetime.now().strftime("%Y-%m-%d")

    for i in range(n_levels):
        lvl_type = (level_types[i] or "").strip()
        lvl_method = (level_methods[i] or "").strip()
        lvl_cur = (level_currencies[i] or "").strip()
        lvl_moq_uom = (level_moq_uoms[i] or "").strip()
        lvl_moq_qty = (level_moq_qtys[i] or "").strip()
        lvl_change = (level_price_change_types[i] or "").strip() or "A"
        row_label = f"Pricing Level {i+1}"

        # Check if row completely blank
        if not (
            lvl_type or lvl_method or lvl_cur or lvl_moq_uom or lvl_moq_qty
            or net_list[i] or net_costs[i] or net_effective_dates[i]
            or pl_list[i] or pl_jobber[i] or pl_net[i] or pl_effective_dates[i]
            or db_base[i] or db_discount[i] or db_effective_dates[i]
            or ehc_base[i] or pr_price[i] or qt_price[i] or td_price[i]
        ):
            continue

        any_level_used = True

        # Level Type
        if not lvl_type:
            errors.append(f"{row_label}: Level Type is required.")
        elif lvl_type not in ["Each", "Case", "Pallet", "Bulk"]:
            errors.append(f"{row_label}: Level Type must be Each, Case, Pallet, or Bulk.")

        # Pricing Method
        if not lvl_method:
            errors.append(f"{row_label}: Pricing Method is required.")
        elif lvl_method not in PRICING_METHODS.keys():
            errors.append(f"{row_label}: Invalid Pricing Method selected.")

        # Currency
        if not lvl_cur:
            errors.append(f"{row_label}: Currency is required.")
        elif lvl_cur not in ["CAD", "USD"]:
            errors.append(f"{row_label}: Currency must be CAD or USD.")

        # MOQ
        if lvl_moq_qty:
            try:
                float(lvl_moq_qty)
            except ValueError:
                errors.append(f"{row_label}: MOQ must be numeric.")
        if lvl_moq_uom and lvl_moq_uom not in QUANTITY_UOM:
            errors.append(f"{row_label}: MOQ Unit must be a valid unit ({', '.join(QUANTITY_UOM)}).")

        # Tier Min / Max
        tmin = (level_tier_min_qtys[i] or "").strip()
        tmax = (level_tier_max_qtys[i] or "").strip()
        for val, label_val in [(tmin, "Tier Min Qty"), (tmax, "Tier Max Qty")]:
            if val:
                try:
                    float(val)
                except ValueError:
                    errors.append(f"{row_label}: {label_val} must be numeric.")

        # Method-specific checks
        if lvl_method == "net_cost":
            nl = (net_list[i] or "").strip()
            nc = (net_costs[i] or "").strip()
            ne = (net_effective_dates[i] or "").strip()

            if not nl:
                errors.append(f"{row_label} (Net Cost): List Price is required.")
            else:
                try:
                    float(nl)
                except ValueError:
                    errors.append(f"{row_label} (Net Cost): List Price must be numeric.")

            if not nc:
                errors.append(f"{row_label} (Net Cost): Net Cost is required.")
            else:
                try:
                    float(nc)
                except ValueError:
                    errors.append(f"{row_label} (Net Cost): Net Cost must be numeric.")

            if not ne:
                errors.append(f"{row_label} (Net Cost): Effective Date is required.")
            elif ne < today_str:
                errors.append(f"{row_label} (Net Cost): Effective Date cannot be before today.")

        elif lvl_method == "price_levels":
            ll = (pl_list[i] or "").strip()
            lj = (pl_jobber[i] or "").strip()
            ln = (pl_net[i] or "").strip()
            le = (pl_effective_dates[i] or "").strip()

            for val, name in [(ll, "List Price"), (lj, "Jobber Price"), (ln, "Net Cost")]:
                if not val:
                    errors.append(f"{row_label} (Price Levels): {name} is required.")
                else:
                    try:
                        float(val)
                    except ValueError:
                        errors.append(f"{row_label} (Price Levels): {name} must be numeric.")

            if not le:
                errors.append(f"{row_label} (Price Levels): Effective Date is required.")
            elif le < today_str:
                errors.append(f"{row_label} (Price Levels): Effective Date cannot be before today.")

        elif lvl_method == "discount_based":
            dbb = (db_base[i] or "").strip()
            dbd = (db_discount[i] or "").strip()
            dbe = (db_effective_dates[i] or "").strip()

            if not dbb:
                errors.append(f"{row_label} (Discount): Base Price is required.")
            else:
                try:
                    float(dbb)
                except ValueError:
                    errors.append(f"{row_label} (Discount): Base Price must be numeric.")

            if not dbd:
                errors.append(f"{row_label} (Discount): Discount % is required.")
            else:
                try:
                    val = float(dbd)
                    if val < 0 or val > 1:
                        errors.append(f"{row_label} (Discount): Discount % must be between 0.0 and 1.0.")
                except ValueError:
                    errors.append(f"{row_label} (Discount): Discount % must be numeric.")

            if not dbe:
                errors.append(f"{row_label} (Discount): Effective Date is required.")
            elif dbe < today_str:
                errors.append(f"{row_label} (Discount): Effective Date cannot be before today.")

        elif lvl_method == "ehc_based":
            eb = (ehc_base[i] or "").strip()
            if not eb:
                errors.append(f"{row_label} (EHC): Base Price is required.")
            else:
                try:
                    float(eb)
                except ValueError:
                    errors.append(f"{row_label} (EHC): Base Price must be numeric.")

            # Numeric checks for EHC fee columns + packaging numbers
            numeric_lists = [
                (ehc_cb[i], "Canadian Blue"),
                (ehc_qty_case[i], "Qty/Case"),
                (ehc_moq[i], "Packaging MOQ"),
                (ehc_abmbsk_each[i], "AB/MB/SK Each"),
                (ehc_abmbsk_case[i], "AB/MB/SK Case"),
                (ehc_bc_each[i], "BC Each"),
                (ehc_bc_case[i], "BC Case"),
                (ehc_nl_each[i], "NL Each"),
                (ehc_nl_case[i], "NL Case"),
                (ehc_ns_each[i], "NS Each"),
                (ehc_ns_case[i], "NS Case"),
                (ehc_nbqc_each[i], "NB/QC Each"),
                (ehc_nbqc_case[i], "NB/QC Case"),
                (ehc_pei_each[i], "PEI Each"),
                (ehc_pei_case[i], "PEI Case"),
                (ehc_yk_each[i], "YK Each"),
                (ehc_yk_case[i], "YK Case"),
            ]
            for val, label_val in numeric_lists:
                v = (val or "").strip()
                if v:
                    try:
                        float(v)
                    except ValueError:
                        errors.append(f"{row_label} (EHC): {label_val} must be numeric.")

        elif lvl_method == "promo_pricing":
            pp = (pr_price[i] or "").strip()
            ps = (pr_start[i] or "").strip()
            pe = (pr_end[i] or "").strip()

            if not pp:
                errors.append(f"{row_label} (Promo): Promo Price is required.")
            else:
                try:
                    float(pp)
                except ValueError:
                    errors.append(f"{row_label} (Promo): Promo Price must be numeric.")

            if not ps:
                errors.append(f"{row_label} (Promo): Start Date is required.")
            if not pe:
                errors.append(f"{row_label} (Promo): End Date is required.")
            if ps and pe and ps > pe:
                errors.append(f"{row_label} (Promo): End Date cannot be before Start Date.")

        elif lvl_method == "quote_pricing":
            qp = (qt_price[i] or "").strip()
            qs = (qt_start[i] or "").strip()
            qe = (qt_end[i] or "").strip()

            if not qp:
                errors.append(f"{row_label} (Quote): Quote Price is required.")
            else:
                try:
                    float(qp)
                except ValueError:
                    errors.append(f"{row_label} (Quote): Quote Price must be numeric.")

            if not qs:
                errors.append(f"{row_label} (Quote): Start Date is required.")
            if not qe:
                errors.append(f"{row_label} (Quote): End Date is required.")
            if qs and qe and qs > qe:
                errors.append(f"{row_label} (Quote): End Date cannot be before Start Date.")

        elif lvl_method == "tender_pricing":
            tp = (td_price[i] or "").strip()
            ts = (td_start[i] or "").strip()
            te = (td_end[i] or "").strip()

            if not tp:
                errors.append(f"{row_label} (Tender): Tender Price is required.")
            else:
                try:
                    float(tp)
                except ValueError:
                    errors.append(f"{row_label} (Tender): Tender Price must be numeric.")

            if not ts:
                errors.append(f"{row_label} (Tender): Start Date is required.")
            if not te:
                errors.append(f"{row_label} (Tender): End Date is required.")
            if ts and te and ts > te:
                errors.append(f"{row_label} (Tender): End Date cannot be before Start Date.")

    if not any_level_used:
        errors.append("At least one valid pricing level must be entered.")

    return (len(errors) == 0, errors)


# ---------------------------------------
# LOOKUP CONSTANTS / CONTROLLED VOCAB
# ---------------------------------------

VENDOR_LIST = [
    # OptiCat vendors
    "Dayton Parts", "Grote Lighting", "Neapco",
    "Truck Lite", "Baldwin Filters", "Stemco", "High Bar Brands",
    # Non-OptiCat vendors
    "Ride Air", "Tetran", "SAF Holland",
    "Consolidated Metco", "Tiger Tool", "J.W Speaker", "Rigid Industries"
]

CHANGE_TYPES = ["A", "M", "D"]  # Add, Modify, Delete

QUANTITY_UOM = ["EA", "PC", "BOX", "CS", "PK", "SET", "RL", "BG", "BT", "DZ"]
PACKAGING_TYPES = ["BOX", "CASE", "PALLET", "BAG", "BOTTLE", "CAN", "CARTON", "WRAP", "PACK", "CRATE", "TUBE"]
PACKAGE_UOM = ["EA", "BX", "CS", "PL"]


WEIGHT_UOM = ["LB", "KG", "G", "OZ"]
DIMENSION_UOM = ["IN", "CM", "MM"]

LIFECYCLE_STATUS = ["Active", "Inactive", "Obsolete"]
HAZMAT_OPTIONS = ["Y", "N"]
APPLICATION_FLAG = ["Y", "N"]

GTIN_TYPES = ["UPC", "EAN"]

DESCRIPTION_TYPES = ["Short", "Long", "Marketing", "Web", "Extended", "Technical"]

LANGUAGES = ["EN", "FR", "ES"]

MEDIA_TYPES = ["MainImage", "AngleImage", "PDF", "SpecSheet", "Logo", "Thumbnail", "InstallationGuide"]
FILE_FORMATS = ["JPEG", "PNG", "GIF", "PDF", "WEBP"]
COLOR_MODES = ["RGB", "CMYK"]
MEDIA_DATE_TYPES = ["Created", "Updated"]

CURRENCIES = ["CAD", "USD"]

ALT_UOM = ["EA", "PC", "SET", "BOX"]

PRICING_METHODS = {
    "net_cost": "Net Cost Provided",
    "price_levels": "Price Levels Provided",
    "discount_based": "Discount-Based Pricing",
    "ehc_based": "EHC-Based Pricing",
    "promo_pricing": "Promo Pricing",
    "quote_pricing": "Quote Pricing",
    "tender_pricing": "Tender Pricing",
}


PRODUCT_STATUS = ["Active", "Inactive", "Obsolete"]



# ---------------------------------------
# ROUTES
# ---------------------------------------

@app.route('/')
def index():
    return redirect(url_for('upload_page'))


@app.route('/upload', methods=['GET'])
def upload_page():
    return render_template('uploads.html')


@app.route('/upload', methods=['POST'])
def upload_files():
    vendor_name = request.form.get('vendor_name')
    zip_path = request.form.get('zip_path')

    # if vendor_name not in OPTICAT_VENDORS:
    #     flash('Only OptiCat vendors are supported in Phase 1.', 'danger')
    #     return redirect(url_for('upload_page'))

    product_file = request.files.get('product_file')
    pricing_file = request.files.get('pricing_file')

    if not product_file or not pricing_file:
        flash('Product XML and Pricing XLSX are required.', 'danger')
        return redirect(url_for('upload_page'))

    product_path = save_file(product_file, vendor_name, "opticat")
    pricing_path = save_file(pricing_file, vendor_name, "opticat")

    try:
        upload_to_azure_bronze(vendor_name, product_path, pricing_path, zip_path)
        flash(f'Files for {vendor_name} uploaded to Azure Bronze successfully.', 'success')
    except Exception as e:
        print(f"❌ Azure upload failed: {e}")
        flash(f'Azure upload failed: {e}', 'danger')

    return redirect(url_for('upload_page'))


@app.route('/download-template')
def download_template():
    try:
        return send_from_directory(TEMPLATE_FOLDER, "standard_template.xlsx", as_attachment=True)
    except FileNotFoundError:
        flash("Template not found on server.", "danger")
        return redirect(url_for('upload_page'))


@app.route('/form-help')
def form_submission_help():
    return "<h4>Coming soon: Vendor submission help guide</h4>"

@app.route('/single-product', methods=['GET'])
def single_product_page():
    pending = session.get('pending_products', [])
    vendor_prefill = session.get('single_vendor_name', "")
    latest_excel = session.get('latest_single_products_excel_path', None)

    return render_template(
        'single_product.html',
        VENDOR_LIST=VENDOR_LIST,
        PRODUCT_STATUS=PRODUCT_STATUS,
        QUANTITY_UOM=QUANTITY_UOM,
        PACKAGE_UOM=PACKAGE_UOM,
        WEIGHT_UOM=WEIGHT_UOM,
        PRICING_METHODS=PRICING_METHODS,
        pending_products=pending,
        vendor_prefill=vendor_prefill,
        latest_excel_available=bool(latest_excel)
    )


@app.route('/single-product', methods=['POST'])
@app.route('/single-product', methods=['POST'])
def submit_single_product():
    action = request.form.get("action")
    ok, errors = validate_single_product_new(request.form)

    vendor_name = request.form.get("vendor_name", "").strip()
    sku = request.form.get("sku", "").strip()
    product_status = request.form.get("product_status", "").strip()
    pricing_method_key = request.form.get("pricing_method", "").strip()

    session['single_vendor_name'] = vendor_name

    if not ok:
        for e in errors:
            flash(e, "danger")
        return redirect(url_for('single_product_page'))

    # --------- Pull / init batch lists from session ----------
    item_rows = session.get("batch_item_rows", [])
    desc_rows = session.get("batch_desc_rows", [])
    ext_rows = session.get("batch_ext_rows", [])
    attr_rows = session.get("batch_attr_rows", [])
    interchange_rows = session.get("batch_interchange_rows", [])
    package_rows = session.get("batch_package_rows", [])
    asset_rows = session.get("batch_asset_rows", [])
    price_rows = session.get("batch_price_rows", [])
    pending = session.get("pending_products", [])

    # --------- SECTION 1: Item Master row ----------
    unspsc = request.form.get("unspsc_code", "").strip()
    hazmat = request.form.get("hazmat_flag", "").strip()
    barcode_type = request.form.get("barcode_type", "").strip()
    barcode_number = request.form.get("barcode_number", "").strip()
    quantity_uom = request.form.get("quantity_uom", "").strip()
    quantity_size = request.form.get("quantity_size", "").strip()
    vmrs = request.form.get("vmrs_code", "").strip()

    item_rows.append({
        "Vendor": vendor_name,
        "Part Number": sku,
        "UNSPSC": unspsc,
        "HazmatFlag": hazmat,
        "Product Status": product_status,
        "Barcode Type": barcode_type,
        "Barcode Number": barcode_number,
        "Quantity UOM": quantity_uom,
        "Quantity Size": quantity_size,
        "VMRS Code": vmrs,
    })

    # --------- SECTION 2: Descriptions (1:M) ----------
    desc_change_types = request.form.getlist("desc_change_type[]")
    desc_codes = request.form.getlist("desc_code[]")
    desc_values = request.form.getlist("desc_value[]")
    desc_sequences = request.form.getlist("desc_sequence[]")

    for ct, code, val, seq in zip(desc_change_types, desc_codes, desc_values, desc_sequences):
        if not (ct or code or val or seq):
            continue
        desc_rows.append({
            "SKU": sku,
            "Description Change Type": ct.strip(),
            "Description Code": code.strip(),
            "Description Value": val.strip(),
            "Sequence": seq.strip(),
        })

    # --------- SECTION 3: Extended Info (1:M) ----------
    ext_change_types = request.form.getlist("ext_change_type[]")
    ext_codes = request.form.getlist("ext_code[]")
    ext_values = request.form.getlist("ext_value[]")

    for ct, code, val in zip(ext_change_types, ext_codes, ext_values):
        if not (ct or code or val):
            continue
        ext_rows.append({
            "SKU": sku,
            "Extended Info Change Type": ct.strip(),
            "Extended Info Code": code.strip(),
            "Extended Info Value": val.strip(),
        })

    # --------- SECTION 4: Attributes (1:M) ----------
    attr_change_types = request.form.getlist("attr_change_type[]")
    attr_names = request.form.getlist("attr_name[]")
    attr_values = request.form.getlist("attr_value[]")

    for ct, name, val in zip(attr_change_types, attr_names, attr_values):
        if not (ct or name or val):
            continue
        attr_rows.append({
            "SKU": sku,
            "Attribute Change Type": ct.strip(),
            "Attribute Name": name.strip(),
            "Attribute Value": val.strip(),
        })

    # --------- SECTION 5: Part Interchange (1:M) ----------
    # NOTE: Your HTML needs inputs:
    #   int_change_type[]
    #   int_brand_label[]
    #   int_part_number[]
    int_change_types = request.form.getlist("int_change_type[]")
    int_brand_labels = request.form.getlist("int_brand_label[]")
    int_part_numbers = request.form.getlist("int_part_number[]")

    for ct, brand, part in zip(int_change_types, int_brand_labels, int_part_numbers):
        if not (ct or brand or part):
            continue
        interchange_rows.append({
            "SKU": sku,
            "Part Interchange Change Type": ct.strip(),
            "Brand Label": brand.strip(),
            "Part Number": part.strip(),
        })

    # --------- SECTION 6: Packages (1:M) ----------
    pack_change_types = request.form.getlist("pack_change_type[]")
    pack_uoms = request.form.getlist("pack_uom[]")
    pack_qty_each = request.form.getlist("pack_qty_each[]")
    pack_weight_uoms = request.form.getlist("pack_weight_uom[]")
    pack_weights = request.form.getlist("pack_weight[]")

    for ct, uom, qty, wuom, wt in zip(
        pack_change_types, pack_uoms, pack_qty_each, pack_weight_uoms, pack_weights
    ):
        if not (ct or uom or qty or wuom or wt):
            continue
        package_rows.append({
            "SKU": sku,
            "Package Change Type": ct.strip(),
            "Package UOM": uom.strip(),
            "Package Quantity of Eaches": qty.strip(),
            "Weight UOM": wuom.strip(),
            "Weight": wt.strip(),
        })

    # --------- SECTION 7: Digital Assets (1:M) ----------
    asset_change_types = request.form.getlist("asset_change_type[]")
    asset_media_types = request.form.getlist("asset_media_type[]")
    asset_files = request.files.getlist("asset_file[]")

    # Save files to uploads/manual_assets/vendor/sku/
    asset_base_dir = os.path.join(app.config['UPLOAD_FOLDER'], "manual_assets", vendor_name, sku)
    os.makedirs(asset_base_dir, exist_ok=True)

    for ct, mtype, file in zip(asset_change_types, asset_media_types, asset_files):
        filename = file.filename if file else ""
        if not (ct or mtype or filename):
            continue
        safe_name = secure_filename(filename)
        local_path = os.path.join(asset_base_dir, safe_name)
        if filename:
            file.save(local_path)
        asset_rows.append({
            "SKU": sku,
            "Digital Change Type": ct.strip(),
            "Media Type": mtype.strip(),
            "FileName": safe_name,
            "FileLocalPath": local_path,
        })

    # --------- SECTION 8: Pricing (1:M) ----------
        # --------- SECTION 8: Pricing (multi-level) ----------
    level_types = request.form.getlist("level_type[]")
    level_price_change_types = request.form.getlist("level_price_change_type[]")
    level_moq_uoms = request.form.getlist("level_moq_uom[]")
    level_moq_qtys = request.form.getlist("level_moq_qty[]")
    level_currencies = request.form.getlist("level_currency[]")
    level_methods = request.form.getlist("level_pricing_method[]")
    level_tier_min_qtys = request.form.getlist("level_tier_min_qty[]")
    level_tier_max_qtys = request.form.getlist("level_tier_max_qty[]")

    net_list = request.form.getlist("level_net_list_price[]")
    net_costs = request.form.getlist("level_net_net_cost[]")
    net_effective_dates = request.form.getlist("level_net_effective_date[]")

    pl_list = request.form.getlist("level_pl_list_price[]")
    pl_jobber = request.form.getlist("level_pl_jobber_price[]")
    pl_net = request.form.getlist("level_pl_net_cost[]")
    pl_effective_dates = request.form.getlist("level_pl_effective_date[]")

    db_base = request.form.getlist("level_db_base_price[]")
    db_discount = request.form.getlist("level_db_discount_pct[]")
    db_list_opt = request.form.getlist("level_db_list_price_opt[]")
    db_effective_dates = request.form.getlist("level_db_effective_date[]")

    ehc_base = request.form.getlist("level_ehc_base_price[]")
    ehc_cb = request.form.getlist("level_ehc_canadian_blue[]")
    ehc_qty_case = request.form.getlist("level_ehc_qty_case[]")
    ehc_upc_each = request.form.getlist("level_ehc_upc_each[]")
    ehc_upc_case = request.form.getlist("level_ehc_upc_case[]")
    ehc_moq = request.form.getlist("level_ehc_moq[]")

    ehc_abmbsk_each = request.form.getlist("level_ehc_abmbsk_each[]")
    ehc_abmbsk_case = request.form.getlist("level_ehc_abmbsk_case[]")
    ehc_bc_each = request.form.getlist("level_ehc_bc_each[]")
    ehc_bc_case = request.form.getlist("level_ehc_bc_case[]")
    ehc_nl_each = request.form.getlist("level_ehc_nl_each[]")
    ehc_nl_case = request.form.getlist("level_ehc_nl_case[]")
    ehc_ns_each = request.form.getlist("level_ehc_ns_each[]")
    ehc_ns_case = request.form.getlist("level_ehc_ns_case[]")
    ehc_nbqc_each = request.form.getlist("level_ehc_nbqc_each[]")
    ehc_nbqc_case = request.form.getlist("level_ehc_nbqc_case[]")
    ehc_pei_each = request.form.getlist("level_ehc_pei_each[]")
    ehc_pei_case = request.form.getlist("level_ehc_pei_case[]")
    ehc_yk_each = request.form.getlist("level_ehc_yk_each[]")
    ehc_yk_case = request.form.getlist("level_ehc_yk_case[]")

    pr_price = request.form.getlist("level_pr_promo_price[]")
    pr_start = request.form.getlist("level_pr_start_date[]")
    pr_end = request.form.getlist("level_pr_end_date[]")

    qt_price = request.form.getlist("level_qt_price[]")
    qt_number = request.form.getlist("level_qt_number[]")
    qt_start = request.form.getlist("level_qt_start_date[]")
    qt_end = request.form.getlist("level_qt_end_date[]")

    td_price = request.form.getlist("level_td_price[]")
    td_number = request.form.getlist("level_td_number[]")
    td_start = request.form.getlist("level_td_start_date[]")
    td_end = request.form.getlist("level_td_end_date[]")

    n_levels = len(level_types)

    def pad(lst):
        return lst + [""] * (n_levels - len(lst))

    level_price_change_types = pad(level_price_change_types)
    level_moq_uoms = pad(level_moq_uoms)
    level_moq_qtys = pad(level_moq_qtys)
    level_currencies = pad(level_currencies)
    level_methods = pad(level_methods)
    level_tier_min_qtys = pad(level_tier_min_qtys)
    level_tier_max_qtys = pad(level_tier_max_qtys)

    net_list = pad(net_list)
    net_costs = pad(net_costs)
    net_effective_dates = pad(net_effective_dates)
    pl_list = pad(pl_list)
    pl_jobber = pad(pl_jobber)
    pl_net = pad(pl_net)
    pl_effective_dates = pad(pl_effective_dates)
    db_base = pad(db_base)
    db_discount = pad(db_discount)
    db_list_opt = pad(db_list_opt)
    db_effective_dates = pad(db_effective_dates)
    ehc_base = pad(ehc_base)
    ehc_cb = pad(ehc_cb)
    ehc_qty_case = pad(ehc_qty_case)
    ehc_upc_each = pad(ehc_upc_each)
    ehc_upc_case = pad(ehc_upc_case)
    ehc_moq = pad(ehc_moq)
    ehc_abmbsk_each = pad(ehc_abmbsk_each)
    ehc_abmbsk_case = pad(ehc_abmbsk_case)
    ehc_bc_each = pad(ehc_bc_each)
    ehc_bc_case = pad(ehc_bc_case)
    ehc_nl_each = pad(ehc_nl_each)
    ehc_nl_case = pad(ehc_nl_case)
    ehc_ns_each = pad(ehc_ns_each)
    ehc_ns_case = pad(ehc_ns_case)
    ehc_nbqc_each = pad(ehc_nbqc_each)
    ehc_nbqc_case = pad(ehc_nbqc_case)
    ehc_pei_each = pad(ehc_pei_each)
    ehc_pei_case = pad(ehc_pei_case)
    ehc_yk_each = pad(ehc_yk_each)
    ehc_yk_case = pad(ehc_yk_case)
    pr_price = pad(pr_price)
    pr_start = pad(pr_start)
    pr_end = pad(pr_end)
    qt_price = pad(qt_price)
    qt_number = pad(qt_number)
    qt_start = pad(qt_start)
    qt_end = pad(qt_end)
    td_price = pad(td_price)
    td_number = pad(td_number)
    td_start = pad(td_start)
    td_end = pad(td_end)

    pricing_method_labels_set = set()

    for i in range(n_levels):
        lvl_type = (level_types[i] or "").strip()
        lvl_change = (level_price_change_types[i] or "").strip() or "A"
        lvl_moq_uom = (level_moq_uoms[i] or "").strip()
        lvl_moq_qty = (level_moq_qtys[i] or "").strip()
        lvl_currency = (level_currencies[i] or "").strip()
        lvl_method = (level_methods[i] or "").strip()

        if not (
            lvl_type or lvl_method or lvl_currency or lvl_moq_uom or lvl_moq_qty
            or net_list[i] or net_costs[i] or net_effective_dates[i]
            or pl_list[i] or pl_jobber[i] or pl_net[i] or pl_effective_dates[i]
            or db_base[i] or db_discount[i] or db_effective_dates[i]
            or ehc_base[i] or pr_price[i] or qt_price[i] or td_price[i]
        ):
            continue

        method_label = PRICING_METHODS.get(lvl_method, lvl_method)
        pricing_method_labels_set.add(method_label)

        row = {
            "Vendor": vendor_name,
            "Part Number": sku,
            "Pricing Method": method_label,
            "Currency": lvl_currency,
            "MOQ Unit": lvl_moq_uom,
            "MOQ": lvl_moq_qty,
            "Pricing Change Type": lvl_change,
            "Pricing Type": lvl_type,  # Each / Case / Pallet / Bulk
            "List Price": "",
            "Jobber Price": "",
            "Discount %": "",
            "Multiplier": "",
            "Pricing Amount": "",
            "Tier Min Qty": (level_tier_min_qtys[i] or "").strip(),
            "Tier Max Qty": (level_tier_max_qtys[i] or "").strip(),
            "Effective Date": "",
            "Start Date": "",
            "End Date": "",
            "Core Part Number": "",
            "Core Cost": "",
            "Notes": "",
        }

        if lvl_method == "net_cost":
            row["List Price"] = (net_list[i] or "").strip()
            row["Pricing Amount"] = row["Net Price"] = (net_costs[i] or "").strip()
            row["Effective Date"] = (net_effective_dates[i] or "").strip()

        elif lvl_method == "price_levels":
            row["List Price"] = (pl_list[i] or "").strip()
            row["Jobber Price"] = (pl_jobber[i] or "").strip()
            row["Pricing Amount"] = row["Net Price"] = (pl_net[i] or "").strip()
            row["Effective Date"] = (pl_effective_dates[i] or "").strip()

        elif lvl_method == "discount_based":
            base_val = (db_base[i] or "").strip()
            disc_val = (db_discount[i] or "").strip()
            row["List Price"] = (db_list_opt[i] or "").strip()
            row["Discount %"] = disc_val
            try:
                b = float(base_val)
                d = float(disc_val) if disc_val else 0.0
                net_val = b * (1 - d)
                row["Pricing Amount"] = row["Net Price"] = f"{net_val:.4f}"
            except ValueError:
                row["Pricing Amount"] = row["Net Price"] = base_val
            row["Effective Date"] = (db_effective_dates[i] or "").strip()

        elif lvl_method == "ehc_based":
            base = (ehc_base[i] or "").strip()
            row["Pricing Amount"] = row["Net Price"] = base
            row["Notes"] = "EHC fees provided by region."

        elif lvl_method == "promo_pricing":
            row["Pricing Amount"] = (pr_price[i] or "").strip()
            row["Start Date"] = (pr_start[i] or "").strip()
            row["End Date"] = (pr_end[i] or "").strip()

        elif lvl_method == "quote_pricing":
            row["Pricing Amount"] = (qt_price[i] or "").strip()
            row["Start Date"] = (qt_start[i] or "").strip()
            row["End Date"] = (qt_end[i] or "").strip()
            qn = (qt_number[i] or "").strip()
            if qn:
                row["Notes"] = f"Quote #: {qn}"

        elif lvl_method == "tender_pricing":
            row["Pricing Amount"] = (td_price[i] or "").strip()
            row["Start Date"] = (td_start[i] or "").strip()
            row["End Date"] = (td_end[i] or "").strip()
            tn = (td_number[i] or "").strip()
            if tn:
                row["Notes"] = f"Tender #: {tn}"

        price_rows.append(row)

    pricing_method_label_summary = (
        ", ".join(sorted(pricing_method_labels_set)) if pricing_method_labels_set else ""
    )


    # --------- Update pending products summary (for UI table) ----------
    pending.append({
        "sku": sku,
        "vendor_name": vendor_name,
        "product_status": product_status,
        "pricing_method": pricing_method_label_summary,
    })

    # --------- Save back to session ----------
    session['batch_item_rows'] = item_rows
    session['batch_desc_rows'] = desc_rows
    session['batch_ext_rows'] = ext_rows
    session['batch_attr_rows'] = attr_rows
    session['batch_interchange_rows'] = interchange_rows
    session['batch_package_rows'] = package_rows
    session['batch_asset_rows'] = asset_rows
    session['batch_price_rows'] = price_rows
    session['pending_products'] = pending

    # --------- Handle action: Add vs Generate ----------
    if action == "add":
        flash(f"Product {sku} added to batch. You can add more or Generate Excel.", "success")
        return redirect(url_for('single_product_page'))

    if action == "generate":
        output_dir = os.path.join(app.config['UPLOAD_FOLDER'], "single_product_batches")
        excel_path = create_multi_product_excel(
            item_rows,
            desc_rows,
            ext_rows,
            attr_rows,
            interchange_rows,
            package_rows,
            asset_rows,
            price_rows,
            output_dir,
        )
        session['latest_single_products_excel_path'] = excel_path

        # Clear batch after generation
        session['batch_item_rows'] = []
        session['batch_desc_rows'] = []
        session['batch_ext_rows'] = []
        session['batch_attr_rows'] = []
        session['batch_interchange_rows'] = []
        session['batch_package_rows'] = []
        session['batch_asset_rows'] = []
        session['batch_price_rows'] = []
        session['pending_products'] = []

        flash(f"Excel generated for {vendor_name}. You can download it below.", "success")
        return redirect(url_for('single_product_page'))

    # Fallback
    return redirect(url_for('single_product_page'))




@app.route('/download-single-products-excel')
def download_single_products_excel():
    excel_path = session.get('latest_single_products_excel_path')
    if not excel_path or not os.path.isfile(excel_path):
        flash("No generated Excel file found for download.", "warning")
        return redirect(url_for('single_product_page'))

    directory, filename = os.path.split(excel_path)
    return send_from_directory(directory, filename, as_attachment=True)




# ---------------------------------------
# MAIN
# ---------------------------------------
if __name__ == '__main__':
    app.run(debug=True)
