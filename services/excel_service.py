"""
Excel file generation services for FGI Vendor Portal
"""
import os
from datetime import datetime
from openpyxl import Workbook


def create_multi_product_excel(
    item_rows,
    desc_rows,
    ext_rows,
    attr_rows,
    interchange_rows,
    package_rows,
    asset_rows,
    price_rows,
    output_dir
):
    """
    Create an Excel file with separate tabs for:
      - Item_Master
      - Descriptions
      - Extended_Info
      - Attributes
      - Part_Interchange
      - Packages
      - Digital_Assets
      - Pricing

    Each *_rows argument is a list of dicts where keys match the column names.
    
    Args:
        item_rows (list): List of item master dictionaries
        desc_rows (list): List of description dictionaries
        ext_rows (list): List of extended info dictionaries
        attr_rows (list): List of attribute dictionaries
        interchange_rows (list): List of part interchange dictionaries
        package_rows (list): List of package dictionaries
        asset_rows (list): List of digital asset dictionaries
        price_rows (list): List of pricing dictionaries
        output_dir (str): Directory path where Excel file will be saved
        
    Returns:
        str: Full filepath of the created Excel file
    """
    os.makedirs(output_dir, exist_ok=True)

    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    filename = f"single_products_batch_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)

    wb = Workbook()

    # 1) Item_Master
    ws = wb.active
    ws.title = "Item_Master"
    item_cols = [
        "Vendor",
        "Part Number",
        "UNSPSC",
        "HazmatFlag",
        "Product Status",
        "Barcode Type",
        "Barcode Number",
        "Quantity UOM",
        "Quantity Size",
        "VMRS Code",
    ]
    ws.append(item_cols)
    for row in item_rows:
        ws.append([
            row.get("Vendor", ""),
            row.get("Part Number", ""),
            row.get("UNSPSC", ""),
            row.get("HazmatFlag", ""),
            row.get("Product Status", ""),
            row.get("Barcode Type", ""),
            row.get("Barcode Number", ""),
            row.get("Quantity UOM", ""),
            row.get("Quantity Size", ""),
            row.get("VMRS Code", ""),
        ])

    # 2) Descriptions
    ws = wb.create_sheet("Descriptions")
    desc_cols = [
        "SKU",
        "Description Change Type",
        "Description Code",
        "Description Value",
        "Sequence",
    ]
    ws.append(desc_cols)
    for row in desc_rows:
        ws.append([
            row.get("SKU", ""),
            row.get("Description Change Type", ""),
            row.get("Description Code", ""),
            row.get("Description Value", ""),
            row.get("Sequence", ""),
        ])

    # 3) Extended_Info
    ws = wb.create_sheet("Extended_Info")
    ext_cols = [
        "SKU",
        "Extended Info Change Type",
        "Extended Info Code",
        "Extended Info Value",
    ]
    ws.append(ext_cols)
    for row in ext_rows:
        ws.append([
            row.get("SKU", ""),
            row.get("Extended Info Change Type", ""),
            row.get("Extended Info Code", ""),
            row.get("Extended Info Value", ""),
        ])

    # 4) Attributes
    ws = wb.create_sheet("Attributes")
    attr_cols = [
        "SKU",
        "Attribute Change Type",
        "Attribute Name",
        "Attribute Value",
    ]
    ws.append(attr_cols)
    for row in attr_rows:
        ws.append([
            row.get("SKU", ""),
            row.get("Attribute Change Type", ""),
            row.get("Attribute Name", ""),
            row.get("Attribute Value", ""),
        ])

    # 5) Part_Interchange
    ws = wb.create_sheet("Part_Interchange")
    interchange_cols = [
        "SKU",
        "Part Interchange Change Type",
        "Brand Label",
        "Part Number",
    ]
    ws.append(interchange_cols)
    for row in interchange_rows:
        ws.append([
            row.get("SKU", ""),
            row.get("Part Interchange Change Type", ""),
            row.get("Brand Label", ""),
            row.get("Part Number", ""),
        ])

    # 6) Packages
    ws = wb.create_sheet("Packages")
    package_cols = [
        "SKU",
        "Package Change Type",
        "Package UOM",
        "Package Quantity of Eaches",
        "Weight UOM",
        "Weight",
    ]
    ws.append(package_cols)
    for row in package_rows:
        ws.append([
            row.get("SKU", ""),
            row.get("Package Change Type", ""),
            row.get("Package UOM", ""),
            row.get("Package Quantity of Eaches", ""),
            row.get("Weight UOM", ""),
            row.get("Weight", ""),
        ])

    # 7) Digital_Assets
    ws = wb.create_sheet("Digital_Assets")
    asset_cols = [
        "SKU",
        "Digital Change Type",
        "Media Type",
        "FileName",
        "FileLocalPath",
    ]
    ws.append(asset_cols)
    for row in asset_rows:
        ws.append([
            row.get("SKU", ""),
            row.get("Digital Change Type", ""),
            row.get("Media Type", ""),
            row.get("FileName", ""),
            row.get("FileLocalPath", ""),
        ])

    # 8) Pricing
    ws = wb.create_sheet("Pricing")
    price_cols = [
    "Vendor", "Part Number", "Pricing Method", "Currency",
    "MOQ Unit", "MOQ", "Pricing Change Type", "Pricing Type",
    "List Price", "Jobber Price", "Discount %", "Multiplier",
    "Pricing Amount",

    # ===== NEW EHC COLUMNS =====
    "EHC AB_MB_SK Each", "EHC AB_MB_SK Case",
    "EHC BC Each", "EHC BC Case",
    "EHC NL Each", "EHC NL Case",
    "EHC NS Each", "EHC NS Case",
    "EHC NB_QC Each", "EHC NB_QC Case",
    "EHC PEI Each", "EHC PEI Case",
    "EHC YK Each", "EHC YK Case",

    "Tier Min Qty", "Tier Max Qty",
    "Effective Date", "Start Date", "End Date",
    "Core Part Number", "Core Cost", "Notes",
]

    ws.append(price_cols)
    for row in price_rows:
        ws.append([
            row.get("Vendor", ""),
            row.get("Part Number", ""),
            row.get("Pricing Method", ""),
            row.get("Currency", ""),
            row.get("MOQ Unit", ""),
            row.get("MOQ", ""),
            row.get("Pricing Change Type", ""),
            row.get("Pricing Type", ""),
            row.get("List Price", ""),
            row.get("Jobber Price", ""),
            row.get("Discount %", ""),
            row.get("Multiplier", ""),
            row.get("Pricing Amount", ""),

            # ==== NEW EHC COLUMNS ====
            row.get("EHC AB_MB_SK Each", ""),
            row.get("EHC AB_MB_SK Case", ""),
            row.get("EHC BC Each", ""),
            row.get("EHC BC Case", ""),
            row.get("EHC NL Each", ""),
            row.get("EHC NL Case", ""),
            row.get("EHC NS Each", ""),
            row.get("EHC NS Case", ""),
            row.get("EHC NB_QC Each", ""),
            row.get("EHC NB_QC Case", ""),
            row.get("EHC PEI Each", ""),
            row.get("EHC PEI Case", ""),
            row.get("EHC YK Each", ""),
            row.get("EHC YK Case", ""),

            # ==== Continue with existing fields ====
            row.get("Tier Min Qty", ""),
            row.get("Tier Max Qty", ""),
            row.get("Effective Date", ""),
            row.get("Start Date", ""),
            row.get("End Date", ""),
            row.get("Core Part Number", ""),
            row.get("Core Cost", ""),
            row.get("Notes", ""),
        ])


    wb.save(filepath)
    return filepath